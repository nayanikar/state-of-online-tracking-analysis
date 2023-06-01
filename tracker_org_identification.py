import os
import json
import openpyxl

def extract_org_name(l_tracker):
	results = {}
	for t in l_tracker:
		for filename in os.listdir('./entities'):
			print(filename)
			if filename.endswith('.json') and t in filename:
				print('found')
				file = './entities/'+filename
				with open(file, 'r') as file:
					data = json.load(file)
					if 'owner' in data:
						owner = data['owner']
						if 'displayName' in owner:
							results[filename] = owner['displayName']
	return results


l_tracker = [
    'livejournal'
]


results = extract_org_name(l_tracker)

print(results)

wb = openpyxl.Workbook()
ws = wb.active

# Write the results to the worksheet
for i, (key, value) in enumerate(results.items(), start=1):
    ws.cell(row=i, column=1, value=key)
    ws.cell(row=i, column=2, value=value)

# Save the workbook to a file
wb.save('results.xlsx')